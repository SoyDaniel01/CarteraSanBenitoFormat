#!/usr/bin/env python3
"""
Módulo de procesamiento para la hoja Cartera_CxC_Det_Comple.
Contiene toda la lógica específica para procesar esta hoja.
"""

from __future__ import annotations

import logging
import math
import unicodedata
from typing import Iterable

import pandas as pd
from openpyxl.utils import get_column_letter, column_index_from_string

logger = logging.getLogger("processor.det_comple")

TARGET_SHEET_NAME = "Cartera_CxC_Det_Comple"
COLUMNS_TO_REMOVE = ["Z", "Y", "X", "O", "M", "L", "K", "I", "D", "C", "B"]
KEYWORDS_TO_CLEAR = {
    "Folio",
    "Fecha",
    "Fecha vencimiento",
    "Condición",
    "Total",
    "Saldo",
    "Dias",
    "No vencido",
    "vencido",
    "30 días",
    "60 días",
    "90 días",
    "120 días",
    "Más de 120 días",
}
SUM_COLUMNS = ["G", "H", "I", "J", "K", "L", "M", "N", "O"]


def normalize_text(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    stripped = "".join(char for char in normalized if not unicodedata.combining(char))
    return stripped.strip().lower()


NORMALIZED_KEYWORDS = {normalize_text(text) for text in KEYWORDS_TO_CLEAR}


def unmerge_all_cells(sheet) -> None:
    merged_ranges = list(sheet.merged_cells.ranges)
    for merged in merged_ranges:
        top_left_value = sheet.cell(merged.min_row, merged.min_col).value
        sheet.unmerge_cells(str(merged))
        for row in range(merged.min_row, merged.max_row + 1):
            for col in range(merged.min_col, merged.max_col + 1):
                cell = sheet.cell(row=row, column=col)
                if cell.value is None:
                    cell.value = top_left_value


def relabel_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [get_column_letter(index + 1) for index in range(df.shape[1])]
    return df


def ensure_min_rows(df: pd.DataFrame, min_rows: int) -> pd.DataFrame:
    if df.shape[0] >= min_rows:
        return df
    missing = min_rows - df.shape[0]
    padding = pd.DataFrame([[pd.NA] * df.shape[1]] * missing, columns=df.columns)
    return pd.concat([df, padding], ignore_index=True)


def ensure_column_span(df: pd.DataFrame, up_to_letter: str) -> pd.DataFrame:
    df = df.copy()
    target_index = column_index_from_string(up_to_letter)
    current_columns = list(df.columns)
    if len(current_columns) < target_index:
        missing = target_index - len(current_columns)
        additional = [
            get_column_letter(len(current_columns) + offset + 1)
            for offset in range(missing)
        ]
        new_columns = current_columns + additional
        df = df.reindex(columns=new_columns)
    return df


def drop_columns_by_letter(df: pd.DataFrame, letters: Iterable[str]) -> pd.DataFrame:
    existing = set(df.columns)
    columns_to_keep = [col for col in df.columns if col not in letters]
    if not columns_to_keep:
        return df
    pruned = df.loc[:, columns_to_keep]
    return relabel_columns(pruned)


def cleanse_keywords(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    def maybe_clear(value, row_index: int):
        if row_index == 4:
            return value
        if isinstance(value, str):
            normalized = normalize_text(value)
            if normalized in NORMALIZED_KEYWORDS:
                return pd.NA
        return value

    cleansed = df.copy()
    for idx in cleansed.index:
        cleansed.loc[idx] = [
            maybe_clear(value, idx) for value in cleansed.loc[idx].tolist()
        ]
    return cleansed


def is_blank_cell(value) -> bool:
    if value is None:
        return True
    if value is pd.NA:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def remove_rows_with_blank_d(df: pd.DataFrame) -> pd.DataFrame:
    if "D" not in df.columns:
        return df

    mask = []
    for idx, value in enumerate(df["D"].tolist()):
        if idx < 5:
            mask.append(True)
            continue
        mask.append(not is_blank_cell(value))

    filtered = df.loc[mask].copy()
    filtered.reset_index(drop=True, inplace=True)
    return filtered


def insert_blank_rows(df: pd.DataFrame, after_row_index: int, rows_to_insert: int) -> pd.DataFrame:
    if rows_to_insert <= 0:
        return df
    top = df.iloc[: after_row_index + 1].copy()
    bottom = df.iloc[after_row_index + 1 :].copy()
    blank_rows = pd.DataFrame(
        [[pd.NA] * df.shape[1]] * rows_to_insert, columns=df.columns
    )
    combined = pd.concat([top, blank_rows, bottom], ignore_index=True)
    return combined


def populate_totals(df: pd.DataFrame) -> pd.DataFrame:
    df = ensure_min_rows(df, 7)
    df = ensure_column_span(df, "O")
    for column in SUM_COLUMNS:
        if column not in df.columns:
            continue
        numeric_series = pd.to_numeric(df.loc[6:, column], errors="coerce")
        total = numeric_series.sum(skipna=True)
        total_value = 0.0 if pd.isna(total) else float(total)
        df.at[4, column] = total_value
    return df


def populate_headers(df: pd.DataFrame) -> pd.DataFrame:
    df = ensure_min_rows(df, 7)
    df = ensure_column_span(df, "O")

    df.at[1, "A"] = "Ferreteria y Madereria San Benito"
    df.at[2, "A"] = "Cartera Detallada completa"
    df.at[0, "G"] = "Cartera vencida"

    total_general = pd.to_numeric(pd.Series([df.at[4, "G"]]), errors="coerce").iloc[0]
    total_vencido = pd.to_numeric(pd.Series([df.at[4, "J"]]), errors="coerce").iloc[0]

    ratio = 0.0
    if not pd.isna(total_general) and not math.isclose(float(total_general), 0.0, abs_tol=1e-9):
        ratio = (float(total_vencido) if not pd.isna(total_vencido) else 0.0) / float(total_general)

    df.at[0, "H"] = ratio * 100
    return df


def forward_fill_column_a_from_row_7(df: pd.DataFrame) -> pd.DataFrame:
    """
    Aplica forward fill en la columna A desde la fila 7 (índice 6) hacia abajo.
    Las celdas vacías toman el valor de la última celda no vacía que esté arriba.
    Revisa celda por celda para manejar todos los tipos de valores vacíos.
    """
    if "A" not in df.columns or df.shape[0] < 7:
        return df
    
    logger.info("Aplicando forward fill en columna A desde fila 7")
    
    # Obtener la serie de la columna A desde la fila 7
    column_a = df["A"].copy()
    last_valid_value = None
    
    # Iterar desde la fila 7 (índice 6) hacia abajo
    for i in range(6, len(column_a)):
        current_value = column_a.iloc[i]
        
        # Verificar si la celda actual está vacía usando la función existente
        if not is_blank_cell(current_value):
            # Si no está vacía, actualizar el último valor válido
            last_valid_value = current_value
            logger.debug(f"Fila {i+1}: Valor válido encontrado: '{current_value}'")
        else:
            # Si está vacía y tenemos un valor válido anterior, copiarlo
            if last_valid_value is not None:
                column_a.iloc[i] = last_valid_value
                logger.debug(f"Fila {i+1}: Copiando valor '{last_valid_value}'")
            else:
                logger.debug(f"Fila {i+1}: Sin valor válido anterior para copiar")
    
    # Actualizar la columna A en el DataFrame
    df["A"] = column_a
    
    logger.info("Forward fill completado en columna A")
    return df


def process_det_comple_sheet(df: pd.DataFrame) -> pd.DataFrame:
    """
    Procesa la hoja Cartera_CxC_Det_Comple con todas las transformaciones específicas.
    """
    logger.info("Iniciando procesamiento de hoja Cartera_CxC_Det_Comple")
    
    if df.empty:
        df = pd.DataFrame(columns=[get_column_letter(i) for i in range(1, 16)])
        df = ensure_min_rows(df, 7)

    df = drop_columns_by_letter(df, COLUMNS_TO_REMOVE)
    df = ensure_column_span(df, "O")
    df = ensure_min_rows(df, 7)

    if "A" in df.columns and df.shape[0] > 3:
        shifted = df.loc[3:, "A"].shift(3)
        df.loc[3:, "A"] = shifted

    df = cleanse_keywords(df)
    df = remove_rows_with_blank_d(df)
    df = ensure_min_rows(df, 7)

    if "A" in df.columns:
        df.loc[4:, "A"] = df.loc[4:, "A"].ffill()

    df.reset_index(drop=True, inplace=True)
    df = insert_blank_rows(df, after_row_index=2, rows_to_insert=2)
    df = ensure_min_rows(df, 7)
    df = ensure_column_span(df, "O")

    df = populate_totals(df)
    df = populate_headers(df)

    df = ensure_min_rows(df, 8)
    if df.shape[0] > 5:
        df = df.drop(index=5).reset_index(drop=True)

    df = ensure_min_rows(df, 7)
    df = ensure_column_span(df, "O")

    if "A" in df.columns:
        df.at[4, "A"] = pd.NA
        df.at[5, "A"] = "Clientes"
    
    # Aplicar forward fill en columna A desde fila 7 hacia abajo
    df = forward_fill_column_a_from_row_7(df)
    
    logger.info("Procesamiento de hoja Cartera_CxC_Det_Comple completado")
    return df


def get_sheet_name() -> str:
    """Retorna el nombre de la hoja que procesa este módulo."""
    return TARGET_SHEET_NAME
