#!/usr/bin/env python3
"""
Script de procesamiento principal para la aplicación Cartera San Benito Format.

Responsabilidades:
- Leer un archivo Excel (.xlsx) desde un path de entrada.
- Transformar la hoja de detalle siguiendo reglas de negocio con pandas.
- Aplicar estilos y crear hojas auxiliares con openpyxl.
- Guardar el resultado en la ruta de salida indicada.
"""

from __future__ import annotations

import logging
import math
import sys
import unicodedata
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string


logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger("processor")

TARGET_SHEET_NAME = "Cartera_CxC_Det_Comple"
ADDITIONAL_SHEETS = [
    "Cartera_CxC_Res_Comp",
    "Cartera_CxC_DSE",
    "Cartera_CxC_RSE",
    "Especiales",
    "Juridico",
    "Proyeccion",
    "Recuperacion",
    "Abonos",
]
COLUMNS_TO_REMOVE = ["Z", "Y", "X", "O", "M", "L", "K", "I", "D", "C", "B"]
KEYWORDS_TO_CLEAR = {
    "Folio",
    "Fecha",
    "Fecha vencimiento",
    "Condición",
    "Total",
    "Saldo",
    "Dias",
    "No vencido",
    "vencido",
    "30 días",
    "60 días",
    "90 días",
    "120 días",
    "Más de 120 días",
}
SUM_COLUMNS = ["G", "H", "I", "J", "K", "L", "M", "N", "O"]
BLUE_FILL = PatternFill(start_color="FF17365D", end_color="FF17365D", fill_type="solid")
WHITE_FONT = Font(color="FFFFFFFF")


def normalize_text(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    stripped = "".join(char for char in normalized if not unicodedata.combining(char))
    return stripped.strip().lower()


NORMALIZED_KEYWORDS = {normalize_text(text) for text in KEYWORDS_TO_CLEAR}


def validate_arguments(args: list[str]) -> tuple[Path, Path]:
    if len(args) < 3:
        raise ValueError("Uso: processor.py <input_path> <output_path>")

    input_path = Path(args[1]).expanduser().resolve()
    output_path = Path(args[2]).expanduser().resolve()

    if not input_path.exists():
        raise FileNotFoundError(f"No se encontró el archivo de entrada: {input_path}")

    if input_path.suffix.lower() != ".xlsx":
        raise ValueError("El archivo de entrada debe tener extensión .xlsx")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    return input_path, output_path


def unmerge_all_cells(sheet) -> None:
    merged_ranges = list(sheet.merged_cells.ranges)
    for merged in merged_ranges:
        top_left_value = sheet.cell(merged.min_row, merged.min_col).value
        sheet.unmerge_cells(str(merged))
        for row in range(merged.min_row, merged.max_row + 1):
            for col in range(merged.min_col, merged.max_col + 1):
                cell = sheet.cell(row=row, column=col)
                if cell.value is None:
                    cell.value = top_left_value


def relabel_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [get_column_letter(index + 1) for index in range(df.shape[1])]
    return df


def ensure_min_rows(df: pd.DataFrame, min_rows: int) -> pd.DataFrame:
    if df.shape[0] >= min_rows:
        return df
    missing = min_rows - df.shape[0]
    padding = pd.DataFrame([[pd.NA] * df.shape[1]] * missing, columns=df.columns)
    return pd.concat([df, padding], ignore_index=True)


def ensure_column_span(df: pd.DataFrame, up_to_letter: str) -> pd.DataFrame:
    df = df.copy()
    target_index = column_index_from_string(up_to_letter)
    current_columns = list(df.columns)
    if len(current_columns) < target_index:
        missing = target_index - len(current_columns)
        additional = [
            get_column_letter(len(current_columns) + offset + 1)
            for offset in range(missing)
        ]
        new_columns = current_columns + additional
        df = df.reindex(columns=new_columns)
    return df


def extract_detail_dataframe(input_path: Path) -> pd.DataFrame:
    workbook = load_workbook(input_path)
    if not workbook.sheetnames:
        workbook.close()
        raise ValueError("El archivo no contiene hojas.")

    worksheet = workbook.active
    worksheet.title = TARGET_SHEET_NAME
    unmerge_all_cells(worksheet)

    max_column = worksheet.max_column
    if max_column == 0:
        workbook.close()
        return pd.DataFrame()

    columns = [get_column_letter(index) for index in range(1, max_column + 1)]
    rows = [
        [cell for cell in row]
        for row in worksheet.iter_rows(min_row=1, max_col=max_column, values_only=True)
    ]
    workbook.close()

    if not rows:
        return pd.DataFrame(columns=columns)

    dataframe = pd.DataFrame(rows, columns=columns)
    dataframe = dataframe.dropna(axis=1, how="all")
    dataframe = relabel_columns(dataframe)
    dataframe = ensure_min_rows(dataframe, 7)
    return dataframe


def drop_columns_by_letter(df: pd.DataFrame, letters: Iterable[str]) -> pd.DataFrame:
    existing = set(df.columns)
    columns_to_keep = [col for col in df.columns if col not in letters]
    if not columns_to_keep:
        # In the unlikely event all columns were removed, keep original data
        return df
    pruned = df.loc[:, columns_to_keep]
    return relabel_columns(pruned)


def cleanse_keywords(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    def maybe_clear(value, row_index: int):
        if row_index == 4:  # preserve row 5
            return value
        if isinstance(value, str):
            normalized = normalize_text(value)
            if normalized in NORMALIZED_KEYWORDS:
                return pd.NA
        return value

    cleansed = df.copy()
    for idx in cleansed.index:
        cleansed.loc[idx] = [
            maybe_clear(value, idx) for value in cleansed.loc[idx].tolist()
        ]
    return cleansed


def is_blank_cell(value) -> bool:
    if value is None:
        return True
    if value is pd.NA:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def remove_rows_with_blank_d(df: pd.DataFrame) -> pd.DataFrame:
    if "D" not in df.columns:
        return df

    mask = []
    for idx, value in enumerate(df["D"].tolist()):
        if idx < 5:
            mask.append(True)
            continue
        mask.append(not is_blank_cell(value))

    filtered = df.loc[mask].copy()
    filtered.reset_index(drop=True, inplace=True)
    return filtered


def insert_blank_rows(df: pd.DataFrame, after_row_index: int, rows_to_insert: int) -> pd.DataFrame:
    if rows_to_insert <= 0:
        return df
    top = df.iloc[: after_row_index + 1].copy()
    bottom = df.iloc[after_row_index + 1 :].copy()
    blank_rows = pd.DataFrame(
        [[pd.NA] * df.shape[1]] * rows_to_insert, columns=df.columns
    )
    combined = pd.concat([top, blank_rows, bottom], ignore_index=True)
    return combined


def populate_totals(df: pd.DataFrame) -> pd.DataFrame:
    df = ensure_min_rows(df, 7)
    df = ensure_column_span(df, "O")
    for column in SUM_COLUMNS:
        if column not in df.columns:
            continue
        numeric_series = pd.to_numeric(df.loc[6:, column], errors="coerce")
        total = numeric_series.sum(skipna=True)
        total_value = 0.0 if pd.isna(total) else float(total)
        df.at[4, column] = total_value
    return df


def populate_headers(df: pd.DataFrame) -> pd.DataFrame:
    df = ensure_min_rows(df, 7)
    df = ensure_column_span(df, "O")

    df.at[1, "A"] = "Ferreteria y Madereria San Benito"
    df.at[2, "A"] = "Cartera Detallada completa"
    df.at[0, "G"] = "Cartera vencida"

    total_general = pd.to_numeric(pd.Series([df.at[4, "G"]]), errors="coerce").iloc[0]
    total_vencido = pd.to_numeric(pd.Series([df.at[4, "J"]]), errors="coerce").iloc[0]

    ratio = 0.0
    if not pd.isna(total_general) and not math.isclose(float(total_general), 0.0, abs_tol=1e-9):
        ratio = (float(total_vencido) if not pd.isna(total_vencido) else 0.0) / float(total_general)

    df.at[0, "H"] = ratio * 100
    return df


def forward_fill_column_a_from_row_7(df: pd.DataFrame) -> pd.DataFrame:
    """
    Aplica forward fill en la columna A desde la fila 7 (índice 6) hacia abajo.
    Las celdas vacías toman el valor de la última celda no vacía que esté arriba.
    Revisa celda por celda para manejar todos los tipos de valores vacíos.
    """
    if "A" not in df.columns or df.shape[0] < 7:
        return df
    
    logger.info("Aplicando forward fill en columna A desde fila 7")
    
    # Obtener la serie de la columna A desde la fila 7
    column_a = df["A"].copy()
    last_valid_value = None
    
    # Iterar desde la fila 7 (índice 6) hacia abajo
    for i in range(6, len(column_a)):
        current_value = column_a.iloc[i]
        
        # Verificar si la celda actual está vacía usando la función existente
        if not is_blank_cell(current_value):
            # Si no está vacía, actualizar el último valor válido
            last_valid_value = current_value
            logger.debug(f"Fila {i+1}: Valor válido encontrado: '{current_value}'")
        else:
            # Si está vacía y tenemos un valor válido anterior, copiarlo
            if last_valid_value is not None:
                column_a.iloc[i] = last_valid_value
                logger.debug(f"Fila {i+1}: Copiando valor '{last_valid_value}'")
            else:
                logger.debug(f"Fila {i+1}: Sin valor válido anterior para copiar")
    
    # Actualizar la columna A en el DataFrame
    df["A"] = column_a
    
    logger.info("Forward fill completado en columna A")
    return df


def process_detail_sheet(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        df = pd.DataFrame(columns=[get_column_letter(i) for i in range(1, 16)])
        df = ensure_min_rows(df, 7)

    df = drop_columns_by_letter(df, COLUMNS_TO_REMOVE)
    df = ensure_column_span(df, "O")
    df = ensure_min_rows(df, 7)

    if "A" in df.columns and df.shape[0] > 3:
        shifted = df.loc[3:, "A"].shift(3)
        df.loc[3:, "A"] = shifted

    df = cleanse_keywords(df)
    df = remove_rows_with_blank_d(df)
    df = ensure_min_rows(df, 7)

    if "A" in df.columns:
        df.loc[4:, "A"] = df.loc[4:, "A"].ffill()

    df.reset_index(drop=True, inplace=True)
    df = insert_blank_rows(df, after_row_index=2, rows_to_insert=2)
    df = ensure_min_rows(df, 7)
    df = ensure_column_span(df, "O")

    df = populate_totals(df)
    df = populate_headers(df)

    df = ensure_min_rows(df, 8)
    if df.shape[0] > 5:
        df = df.drop(index=5).reset_index(drop=True)

    df = ensure_min_rows(df, 7)
    df = ensure_column_span(df, "O")

    if "A" in df.columns:
        df.at[4, "A"] = pd.NA
        df.at[5, "A"] = "Clientes"
    
    # Aplicar forward fill en columna A desde fila 7 hacia abajo
    df = forward_fill_column_a_from_row_7(df)
    
    return df


def dataframe_to_excel(df: pd.DataFrame, output_path: Path) -> None:
    logger.info("Escribiendo archivo procesado en: %s", output_path)
    export_df = df.where(pd.notna(df), None)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        export_df.to_excel(
            writer,
            sheet_name=TARGET_SHEET_NAME,
            index=False,
            header=False,
        )


def apply_styles(output_path: Path) -> None:
    workbook = load_workbook(output_path)
    worksheet = workbook[TARGET_SHEET_NAME]

    for sheet_name in ADDITIONAL_SHEETS:
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(title=sheet_name)

    max_column = worksheet.max_column or column_index_from_string("O")

    for row_index in (5, 6):
        for column_index in range(1, max_column + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            cell.fill = BLUE_FILL
            cell.font = WHITE_FONT

    header_cell = worksheet["G1"]
    header_cell.fill = BLUE_FILL
    header_cell.font = WHITE_FONT

    workbook.save(output_path)
    workbook.close()


def process_workbook(input_path: Path, output_path: Path) -> None:
    logger.info("Procesando archivo: %s", input_path)
    detail_df = extract_detail_dataframe(input_path)
    processed_df = process_detail_sheet(detail_df)
    dataframe_to_excel(processed_df, output_path)
    apply_styles(output_path)
    logger.info("Archivo procesado y guardado en: %s", output_path)


def main() -> int:
    try:
        input_path, output_path = validate_arguments(sys.argv)
        process_workbook(input_path, output_path)
        return 0
    except Exception as exc:  # pylint: disable=broad-except
        logger.error("Error durante el procesamiento: %s", exc)
        return 1


if __name__ == "__main__":
    sys.exit(main())
